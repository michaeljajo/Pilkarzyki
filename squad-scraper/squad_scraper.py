#!/usr/bin/env python3
"""
squad_scraper.py — squad lists for 10 European leagues.

Roster source   : Sofascore (registered squad)
Position source : Flashscore (authoritative), Sofascore fallback
Names           : ASCII, natural order (first name + surname)
Output          : one sheet, four columns — Name, Position, Club, Country of the club
Positions       : Defender / Midfielder / Forward only. Goalkeepers are excluded.

Run in stages so each step is reviewable and resumable:

    python3 squad_scraper.py clubs       # 1. discover clubs per league (Sofascore)
    python3 squad_scraper.py crosswalk   # 2. match clubs to Flashscore -> crosswalk.csv  [REVIEW THIS]
    python3 squad_scraper.py squads      # 3. pull rosters + Flashscore positions
    python3 squad_scraper.py build       # 4. emit squads.xlsx + review.csv

    python3 squad_scraper.py all         # everything end to end
    python3 squad_scraper.py all --pilot # 2 clubs only, for a fast sanity check

Responses are cached so an interrupted run resumes cheaply. The cache lives in
the OS cache directory (~/Library/Caches/squad-scraper on macOS), not next to
this script -- see default_cache_dir(). Override with --cache-dir or the
SQUAD_SCRAPER_CACHE environment variable.

Cached entries expire after 24h by default; --refresh forces a live scrape and
--max-age HOURS sets a different window. The run summary reports how many
responses were fetched versus reused and how large the cache has grown, so
stale output is never silent.

Requires: pip3 install requests openpyxl
"""

from __future__ import annotations

import argparse
import json
import os
import random
import re
import sys
import time
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any
from urllib.parse import quote

# Sofascore sits behind Cloudflare, which fingerprints the TLS handshake (JA3)
# as well as the headers. Stock `requests` on macOS system Python is linked
# against LibreSSL and gets a 403 before the request is even looked at.
# curl_cffi replays a real Chrome handshake and sails through.
try:
    from curl_cffi import requests as _http_lib
    IMPERSONATE = True
except ImportError:
    IMPERSONATE = False
    try:
        import requests as _http_lib
    except ImportError:
        sys.exit("Missing dependency. Run:  pip3 install curl_cffi requests openpyxl")


# ----------------------------------------------------------------------------
# Configuration
# ----------------------------------------------------------------------------

HERE = Path(__file__).resolve().parent
LEGACY_CACHE = HERE / ".cache"
OUT_XLSX = HERE / "squads.xlsx"
OUT_REVIEW = HERE / "review.csv"
OUT_CROSSWALK = HERE / "register.csv"
OUT_CLUBS = HERE / "clubs.json"
OUT_SQUADS = HERE / "squads.json"


def default_cache_dir() -> Path:
    """Where to keep cached HTTP responses.

    Deliberately outside the source tree. This project lives in iCloud Drive,
    and a single full run leaves ~190MB of raw Sofascore JSON behind -- a
    squad endpoint returns three-quarters of a megabyte per club. Kept beside
    the script that is 860 disposable files syncing to every machine. The OS
    cache directory is excluded from backup and sync and gets purged under
    disk pressure, which is the lifetime these files actually deserve.
    """
    override = os.environ.get("SQUAD_SCRAPER_CACHE")
    if override:
        return Path(override).expanduser()
    if sys.platform == "darwin":
        return Path.home() / "Library" / "Caches" / "squad-scraper"
    xdg = os.environ.get("XDG_CACHE_HOME")
    return (Path(xdg).expanduser() if xdg else Path.home() / ".cache") / "squad-scraper"


def dir_size_mb(path: Path) -> float:
    return sum(f.stat().st_size for f in path.glob("*.txt")) / 1e6

# Sofascore unique-tournament IDs are used as a *hint* only. The script verifies
# the tournament name that comes back and falls back to search if it disagrees,
# so a stale ID here degrades to a slower path rather than silently wrong data.
# fs_country is the country slug as it appears in flashscore.pl URLs. It is the
# single most valuable field here: a squad page states its own country, so a
# proposed mapping can be *proved* rather than trusted on name similarity.
LEAGUES: list[dict[str, Any]] = [
    {"label": "Premier League",  "country": "England",     "hint": 17,  "search": "Premier League",         "fs_country": "anglia"},
    {"label": "LaLiga",          "country": "Spain",       "hint": 8,   "search": "LaLiga",                 "fs_country": "hiszpania"},
    {"label": "Ligue 1",         "country": "France",      "hint": 34,  "search": "Ligue 1",                "fs_country": "francja"},
    {"label": "Liga Portugal",   "country": "Portugal",    "hint": 238, "search": "Liga Portugal Betclic",  "fs_country": "portugalia"},
    {"label": "Serie A",         "country": "Italy",       "hint": 23,  "search": "Serie A",                "fs_country": "wlochy"},
    {"label": "Pro League",      "country": "Belgium",     "hint": 38,  "search": "Jupiler Pro League",     "fs_country": "belgia"},
    {"label": "Eredivisie",      "country": "Netherlands", "hint": 37,  "search": "Eredivisie",             "fs_country": "holandia"},
    {"label": "Ekstraklasa",     "country": "Poland",      "hint": 202, "search": "Ekstraklasa",            "fs_country": "polska"},
    {"label": "Super Lig",       "country": "Turkey",      "hint": 52,  "search": "Trendyol Super Lig",     "fs_country": "turcja"},
    {"label": "Bundesliga",      "country": "Germany",     "hint": 35,  "search": "Bundesliga",             "fs_country": "niemcy"},
]

SOFA = "https://api.sofascore.com/api/v1"
FS_BASE = "https://www.flashscore.pl"
FS_SEARCH = "https://s.flashscore.com/search/"
# Read out of flashscore.pl's own page config. Using the wrong language here
# returns slugs for a different edition of the site, which then 404.
FS_LANG, FS_PID = 45, 3

UA = ("Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 "
      "(KHTML, like Gecko) Chrome/124.0 Safari/537.36")

MIN_DELAY = 1.1          # seconds between requests, plus jitter
MAX_RETRIES = 4

# Flashscore position headings, Polish and English, mapped to canonical labels.
FS_HEADINGS = {
    "bramkarze": "Goalkeeper", "goalkeepers": "Goalkeeper",
    "obroncy": "Defender", "defenders": "Defender",
    "pomocnicy": "Midfielder", "midfielders": "Midfielder",
    "napastnicy": "Forward", "forwards": "Forward",
    "trener": "__coach__", "coach": "__coach__", "coaches": "__coach__",
}
SOFA_POS = {"G": "Goalkeeper", "D": "Defender", "M": "Midfielder", "F": "Forward"}

# Output language. Positions and club countries are translated at write time,
# so all the matching logic upstream keeps working on the English values.
# Set POLISH_OUTPUT = False to emit English instead.
POLISH_OUTPUT = True

POSITION_PL = {
    "Defender": "Obrońca",
    "Midfielder": "Pomocnik",
    "Forward": "Napastnik",
    "Goalkeeper": "Bramkarz",     # excluded from output, kept for completeness
}

COUNTRY_PL = {
    "England": "Anglia",
    "Spain": "Hiszpania",
    "France": "Francja",
    "Portugal": "Portugalia",
    "Italy": "Włochy",
    "Belgium": "Belgia",
    "Netherlands": "Holandia",
    "Poland": "Polska",
    "Turkey": "Turcja",
    "Germany": "Niemcy",
}

HEADERS_PL = {
    "Name": "Zawodnik",
    "Position": "Pozycja",
    "Club": "Klub",
    "Country of the club": "Kraj klubu",
}


# ----------------------------------------------------------------------------
# HTTP with cache, rate limiting and retries
# ----------------------------------------------------------------------------

def site_headers(url: str) -> dict[str, str]:
    """Cloudflare also checks that headers look consistent with a real page load."""
    common = {
        "User-Agent": UA,
        "Accept-Language": "en-GB,en;q=0.9,pl;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "sec-ch-ua": '"Chromium";v="124", "Google Chrome";v="124", "Not-A.Brand";v="99"',
        "sec-ch-ua-mobile": "?0",
        "sec-ch-ua-platform": '"macOS"',
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-site",
    }
    if "sofascore" in url:
        common |= {
            "Accept": "*/*",
            "Origin": "https://www.sofascore.com",
            "Referer": "https://www.sofascore.com/",
        }
    else:
        common |= {
            "Accept": "*/*",
            "Origin": FS_BASE,
            "Referer": FS_BASE + "/",
        }
    return common


class Http:
    def __init__(self, verbose: bool = True, max_age_hours: float = 24.0,
                 cache_dir: Path | None = None):
        if IMPERSONATE:
            self.s = _http_lib.Session(impersonate="chrome")
        else:
            self.s = _http_lib.Session()
        self._last = 0.0
        self.verbose = verbose
        # Cached responses older than this are re-fetched. Without an age limit
        # a re-run silently reports yesterday's squads — transfers would never
        # show up. 0 means always refetch.
        self.max_age = max_age_hours * 3600
        self.hits = 0
        self.fetches = 0
        self.cache = Path(cache_dir).expanduser() if cache_dir else default_cache_dir()
        self.cache.mkdir(parents=True, exist_ok=True)
        if verbose and LEGACY_CACHE.is_dir():
            print(f"  ! Leftover in-tree cache at {LEGACY_CACHE} is no longer used.")
            print(f"    Reclaim the space with:  rm -rf '{LEGACY_CACHE}'\n")
        if verbose and not IMPERSONATE:
            print("  ! curl_cffi not installed — Sofascore will likely return 403.")
            print("    Fix with:  pip3 install curl_cffi\n")

    def _throttle(self) -> None:
        wait = MIN_DELAY + random.uniform(0, 0.5) - (time.time() - self._last)
        if wait > 0:
            time.sleep(wait)
        self._last = time.time()

    def get(self, url: str, *, referer: str | None = None, cache_key: str | None = None) -> str:
        key = cache_key or re.sub(r"[^A-Za-z0-9]+", "_", url)[:150]
        path = self.cache / f"{key}.txt"
        if path.exists() and (time.time() - path.stat().st_mtime) < self.max_age:
            self.hits += 1
            return path.read_text(encoding="utf-8")

        headers = site_headers(url)
        if referer:
            headers["Referer"] = referer
        last_err: Exception | None = None
        for attempt in range(MAX_RETRIES):
            self._throttle()
            try:
                r = self.s.get(url, headers=headers, timeout=30)
                if r.status_code == 200:
                    path.write_text(r.text, encoding="utf-8")
                    self.fetches += 1
                    if self.verbose:
                        print(f"    fetched {url[:88]}")
                    return r.text
                if r.status_code in (403, 429, 500, 502, 503):
                    back = 2 ** attempt + random.uniform(0, 1.5)
                    if self.verbose:
                        print(f"    HTTP {r.status_code}, backing off {back:.1f}s")
                    time.sleep(back)
                    last_err = RuntimeError(f"HTTP {r.status_code}")
                    continue
                raise RuntimeError(f"HTTP {r.status_code} for {url}")
            except Exception as e:
                if isinstance(e, RuntimeError) and "HTTP" in str(e) and "403" not in str(e):
                    raise
                last_err = e
                time.sleep(2 ** attempt)
        raise RuntimeError(f"Failed after {MAX_RETRIES} attempts: {url} ({last_err})")

    def get_json(self, url: str, **kw) -> Any:
        return json.loads(self.get(url, **kw))


# ----------------------------------------------------------------------------
# Name normalisation and matching
# ----------------------------------------------------------------------------

# Characters that NFKD does not decompose on its own.
_SPECIAL = str.maketrans({
    "ø": "o", "Ø": "O", "đ": "d", "Đ": "D", "ð": "d", "Ð": "D",
    "ł": "l", "Ł": "L", "æ": "ae", "Æ": "Ae", "œ": "oe", "Œ": "Oe",
    "ß": "ss", "þ": "th", "Þ": "Th", "ı": "i", "İ": "I",
})


def to_ascii(text: str) -> str:
    """Fold to plain ASCII. 'Kerem Aktürkoğlu' -> 'Kerem Akturkoglu'."""
    if not text:
        return ""
    text = text.translate(_SPECIAL)
    text = unicodedata.normalize("NFKD", text)
    text = "".join(c for c in text if not unicodedata.combining(c))
    return text.encode("ascii", "ignore").decode("ascii").strip()


def tokens(name: str) -> frozenset[str]:
    """Order-invariant token set. Kills the surname-first problem outright."""
    cleaned = re.sub(r"[^a-z0-9 ]+", "", to_ascii(name).lower())
    return frozenset(t for t in cleaned.split() if len(t) > 1)


def strip_tags(html: str) -> str:
    return re.sub(r"\s+", " ", re.sub(r"<[^>]+>", " ", html)).strip()


# ----------------------------------------------------------------------------
# Stage 1 — clubs per league, from Sofascore
# ----------------------------------------------------------------------------

def resolve_tournament(http: Http, league: dict) -> int:
    """Verify the hinted tournament ID; fall back to search if it looks wrong."""
    hint = league["hint"]
    try:
        data = http.get_json(f"{SOFA}/unique-tournament/{hint}")
        got = data["uniqueTournament"]["name"]
        want = league["search"]
        if tokens(got) & tokens(want):
            return hint
        print(f"    hint {hint} resolved to '{got}', expected '{want}' — searching")
    except Exception as e:
        print(f"    hint {hint} failed ({e}) — searching")

    res = http.get_json(f"{SOFA}/search/all?q={quote(league['search'])}")
    for item in res.get("results", []):
        if item.get("type") != "uniqueTournament":
            continue
        ent = item["entity"]
        if ent.get("category", {}).get("name", "").lower().startswith(league["country"][:4].lower()):
            return ent["id"]
    for item in res.get("results", []):
        if item.get("type") == "uniqueTournament":
            return item["entity"]["id"]
    raise RuntimeError(f"Could not resolve tournament for {league['label']}")


def current_season(http: Http, tid: int) -> tuple[int, str]:
    seasons = http.get_json(f"{SOFA}/unique-tournament/{tid}/seasons")["seasons"]
    s = seasons[0]
    return s["id"], s.get("year") or s.get("name") or "?"


def league_clubs(http: Http, league: dict) -> list[dict]:
    tid = resolve_tournament(http, league)
    sid, season_label = current_season(http, tid)
    # Printed so a wrong season is obvious immediately. If this ever shows an
    # unexpected year, every club in the league is wrong together.
    print(f"    tournament {tid}, season {season_label}")
    url = f"{SOFA}/unique-tournament/{tid}/season/{sid}/standings/total"
    data = http.get_json(url)

    clubs, seen = [], set()
    for standing in data.get("standings", []):
        for row in standing.get("rows", []):
            team = row.get("team", {})
            if team.get("id") in seen or not team.get("id"):
                continue
            seen.add(team["id"])
            clubs.append({
                "sofa_id": team["id"],
                "name": to_ascii(team.get("name", "")),
                "name_raw": team.get("name", ""),
                "league": league["label"],
                "country": league["country"],
            })
    return clubs


def stage_clubs(http: Http, pilot: bool) -> list[dict]:
    leagues = LEAGUES[:1] if pilot else LEAGUES
    all_clubs: list[dict] = []
    for lg in leagues:
        print(f"  {lg['label']} ({lg['country']})")
        try:
            clubs = league_clubs(http, lg)
        except Exception as e:
            print(f"    !! failed: {e}")
            continue
        if pilot:
            clubs = clubs[:2]
        print(f"    {len(clubs)} clubs")
        all_clubs.extend(clubs)
    OUT_CLUBS.write_text(json.dumps(all_clubs, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  -> {OUT_CLUBS.name} ({len(all_clubs)} clubs)")
    return all_clubs


# ----------------------------------------------------------------------------
# Stage 2 — crosswalk Sofascore club -> Flashscore club
# ----------------------------------------------------------------------------

# Corporate noise in club names that carries no identifying information.
CLUB_STOP = {
    "fc", "cf", "sc", "ac", "as", "ss", "ssc", "afc", "cd", "ud", "rc", "rcd",
    "sd", "cp", "fk", "sk", "nk", "hnk", "bk", "if", "sv", "tsv", "vfl", "vfb",
    "bsc", "kv", "kaa", "kvc", "rsc", "rwd", "ogc", "losc", "us", "club",
    "calcio", "sad", "sa", "team", "kks", "mks", "gks", "ks", "wks", "zks",
}
CLUB_ALIAS = {"utd": "united", "st": "saint", "athl": "athletic", "atl": "atletico"}

# Youth, reserve and women's sides that must never be mistaken for the first team.
# Words that mean "not the first team" wherever they appear in the name.
_MARK_ANY = re.compile(
    r"(?:^|[\s\-.])(u\s?\d{1,2}|jong|beloften|youth|jun(?:iors?)?|acad(?:emy)?|"
    r"amat(?:eure?)?|castilla|uam|women|ladies|kobiety|fem(?:enino)?|frauen|"
    r"feminin\w*|reserves?)(?:$|[\s\-.])", re.I)

# Single letters and numerals only mean it at the END of a name: "Barcelona B",
# "Koln II", "Kolonia K" (K = kobiety on the Polish site). At the start they are
# ordinary abbreviations — "B. Monchengladbach" is Borussia, not a B team.
_MARK_SUFFIX = re.compile(r"[\s\-.](ii|iii|b|c|k|w|m|res)\.?\s*$", re.I)


def not_first_team_marker(name: str) -> str | None:
    """Return the marker showing this is not a first team, or None."""
    for rx in (_MARK_ANY, _MARK_SUFFIX):
        m = rx.search(name or "")
        if m:
            return m.group(1).lower()
    return None


class _Compat:
    """Kept so existing call sites read naturally."""
    @staticmethod
    def search(name: str):
        return not_first_team_marker(name)


NOT_FIRST_TEAM = _Compat()


def club_key(name: str) -> frozenset[str]:
    """Token set for club matching, with corporate suffixes and aliases resolved."""
    raw = re.sub(r"[^a-z0-9 ]+", " ", to_ascii(name).lower()).split()
    out = set()
    for t in raw:
        t = CLUB_ALIAS.get(t, t)
        if t in CLUB_STOP or t.isdigit() or len(t) < 2:
            continue
        out.add(t)
    return frozenset(out)


def fs_candidates(http: Http, name: str) -> list[tuple[str, str, str]]:
    """Return [(flashscore_id, slug, title)] ranked best-first. May be empty."""
    q = quote(name)
    url = f"{FS_SEARCH}?q={q}&l={FS_LANG}&s=1&f=1%3B1&pid={FS_PID}&sid=1"
    try:
        # Language belongs in the cache key: the previous run cached Lithuanian
        # results under a bare name, and reusing those would resurrect the bug.
        raw = http.get(url, referer=FS_BASE,
                       cache_key=f"fssearch_{FS_LANG}_{re.sub(r'[^A-Za-z0-9]+','_',name)}")
    except Exception:
        return []

    # Response is JSONP: cjs.search.jsonpCallback({...}). Unwrap it.
    start, end = raw.find("("), raw.rfind(")")
    body = raw[start + 1:end] if 0 <= start < end else raw
    try:
        data = json.loads(body)
    except json.JSONDecodeError:
        return []

    results = data.get("results") if isinstance(data, dict) else data
    if not isinstance(results, list):
        return []

    want_full = tokens(name)          # unmodified, keeps FC/SC/AC
    want_core = club_key(name)        # suffixes stripped
    scored = []
    for rank, item in enumerate(results):
        if not isinstance(item, dict):
            continue
        # 'type' is the string "participants"; football only.
        if item.get("type") != "participants" or item.get("sport_id") != 1:
            continue

        title = item.get("title") or ""
        # Titles carry a country suffix: "Bournemouth (Anglia)".
        title = re.sub(r"\s*\([^)]*\)\s*$", "", title).strip()
        cand_id, cand_slug = item.get("id"), item.get("url") or ""
        if not cand_id or not title:
            continue

        ct_full, ct_core = tokens(title), club_key(title)
        if not ct_core:
            continue
        # Score on the FULL token set first. Scoring only on the stripped form
        # made "Barcelona SC" identical to "FC Barcelona" — both collapse to
        # {barcelona} — which is how an Ecuadorian club got into LaLiga.
        full = len(want_full & ct_full) / max(len(want_full | ct_full), 1)
        core = len(want_core & ct_core) / max(len(want_core | ct_core), 1)
        score = full + 0.3 * core
        if want_full == ct_full:
            score += 1.0
        elif want_core == ct_core:
            score += 0.3                        # weaker: suffixes were discarded
        # Reserve/youth sides are demoted rather than discarded — verification
        # will reject them on squad overlap, and discarding here would also
        # throw away clubs whose real name contains 'II' or 'B'.
        if NOT_FIRST_TEAM.search(title):
            score -= 0.8
        score -= rank * 0.01
        # Deliberately permissive. Verification (country + squad overlap) is
        # strong enough to reject anything wrong, so a loose net here costs a
        # few extra page fetches but catches clubs the two sites name
        # completely differently.
        if score >= 0.12:
            scored.append((score, str(cand_id), str(cand_slug), title))

    scored.sort(reverse=True)
    return [(cid, slug, title) for _, cid, slug, title in scored[:12]]


def fs_verify(http: Http, fs_id: str, fs_slug: str, want_country: str,
              club_name: str, roster: list[dict]) -> tuple[bool, str]:
    """
    Confirm a candidate really is this club's first team.

    Name and country are necessary but nowhere near sufficient: Fortuna Koln is
    genuinely German and Barcelona Femeni is genuinely Spanish, yet both are the
    wrong club. The decisive test is the squad itself — we already know from
    Sofascore who plays here, so the right Flashscore page must share players
    with it. A women's, reserve or same-city club shares essentially none.

    Returns (ok, reason). The fetched page is cached and reused by stage 3.
    """
    slug = (fs_slug or "team").strip("/").split("/")[-1]
    url = f"{FS_BASE}/druzyna/{slug}/{fs_id}/sklad/"
    try:
        html = http.get(url, referer=FS_BASE, cache_key=f"fssquad_{fs_id}")
    except Exception as e:
        return False, f"unreachable ({str(e)[:40]})"

    m = re.search(r"/pilka-nozna/([a-z0-9\-]+)/", html)
    got_country = m.group(1) if m else ""
    if got_country and got_country != want_country:
        return False, f"country is '{got_country}', expected '{want_country}'"

    title = ""
    tm = re.search(r"<title>(.*?)</title>", html, re.S | re.I)
    if tm:
        title = re.sub(r"^.*?:\s*", "", strip_tags(tm.group(1)))
        title = re.sub(r"\s*-\s*(sk[lł]ad|squad).*$", "", title, flags=re.I).strip()

    # A marker only disqualifies if the club's own name does not contain it.
    # 'Willem II' legitimately carries 'II'; 'Jong Ajax' and 'Kolonia K' do not.
    mark = not_first_team_marker(title) if title else None
    if mark and mark not in tokens(club_name):
        return False, f"not a first team ('{title}', marker '{mark}')"

    # Deliberately NO name comparison here. flashscore.pl localises club names
    # — Marseille is 'Marsylia', Vitoria SC is 'Guimaraes' — so matching on the
    # name rejects correct clubs. The squad overlap below settles identity.
    fs_names = fs_player_tokens(html)
    if not fs_names:
        # Distinguish a data gap from a mismatch: if country and team type are
        # fine, this is probably the right club with no squad published.
        return False, (f"Flashscore lists no squad for '{title or slug}' "
                       f"(likely a data gap, not a wrong match)")

    roster_tokens = [tokens(p["name"]) for p in roster]
    roster_tokens = [t for t in roster_tokens if t]
    if not roster_tokens:
        return False, "no Sofascore roster to check against"

    hits = sum(1 for rt in roster_tokens
               if any(len(rt & ft) >= 2 or rt == ft for ft in fs_names))
    ratio = hits / min(len(roster_tokens), len(fs_names))
    if hits < 5 or ratio < 0.3:
        return False, (f"squad mismatch: only {hits} of {len(roster_tokens)} "
                       f"players found on '{title or slug}'")
    return True, f"verified ({hits} players matched)"


REGISTER_COLS = ["league", "country", "club", "sofa_id", "fs_id", "fs_slug",
                 "status", "checked_by", "note", "lookup"]

# Bump when the verification logic gets stricter. Rows verified under an older,
# weaker rule are re-checked rather than trusted forever — otherwise a bad
# mapping accepted once would survive every future run.
VERIFY_VERSION = "v5-marker-position"

# Clubs the two sites name so differently that one search string cannot find
# both. These are only search *hints* — every hit still has to pass country and
# squad-overlap verification, so a wrong alias cannot inject a wrong club.
SEARCH_ALIASES: dict[str, list[str]] = {
    "Athletic Club": ["Athletic Bilbao", "Bilbao"],
    "Deportivo de A Coruna": ["Deportivo La Coruna", "Deportivo"],
    "Olympique Lyonnais": ["Lyon"],
    "Olympique de Marseille": ["Marsylia", "Marseille"],
    "Paris Saint-Germain": ["PSG", "Paris SG"],
    "Stade Brestois": ["Brest"],
    "Stade Rennais": ["Rennes"],
    "Sint-Truidense VV": ["St. Truiden", "Sint-Truiden"],
    "Willem II Tilburg": ["Willem II"],
    "Amed Sportif Faaliyetler": ["Amedspor", "Amed SK"],
    "Genclerbirligi": ["Genclerbirligi Ankara"],
    "1. FC Koln": ["FC Koln", "Kolonia"],
    # Flashscore abbreviates this one to "B. Monchengladbach", which shares no
    # token at all with "Borussia M'gladbach".
    "Borussia M'gladbach": ["B. Monchengladbach", "Borussia Monchengladbach",
                            "Moenchengladbach"],
    "Vitoria SC": ["Vitoria Guimaraes", "Guimaraes"],
    "FC Bayern Munchen": ["Bayern Monachium", "Bayern Munich"],
    "Inter": ["Inter Mediolan"],
    "AC Milan": ["Milan"],
    "Real Racing Club": ["Racing Santander"],
}


def stage_crosswalk(http: Http, clubs: list[dict]) -> list[dict]:
    """
    Build/refresh register.csv — the durable club register.

    Rows already marked 'verified' are never touched, so hand corrections
    survive every future run. Anything unproven is marked TODO with a ready
    made lookup URL, rather than being force-matched to a plausible wrong club.
    """
    import csv
    fs_country = {lg["label"]: lg["fs_country"] for lg in LEAGUES}

    existing: dict[str, dict] = {}
    if OUT_CROSSWALK.exists():
        with OUT_CROSSWALK.open(encoding="utf-8") as f:
            for r in csv.DictReader(f):
                existing[str(r.get("sofa_id", ""))] = r

    rows, verified, todo = [], 0, 0
    for c in clubs:
        want_country = fs_country.get(c["league"], "")
        prev = existing.get(str(c["sofa_id"]))

        # Trust the register only if it was checked by the current rules, or a
        # human vouched for it explicitly (checked_by=manual).
        if (prev and prev.get("fs_id") and prev.get("status") == "verified"
                and prev.get("checked_by") in (VERIFY_VERSION, "manual")):
            rows.append({**prev, "league": c["league"], "country": c["country"],
                         "club": c["name"], "sofa_id": c["sofa_id"]})
            verified += 1
            continue

        # The roster is the yardstick every candidate is measured against.
        try:
            roster = sofa_roster(http, int(c["sofa_id"]))
        except Exception as e:
            roster = []
            print(f"    !! roster unavailable for {c['name']}: {e}")

        # A hand-entered row is tried first, then search candidates.
        pool: list[tuple[str, str, str]] = []
        if prev and prev.get("fs_id"):
            pool.append((prev["fs_id"], prev.get("fs_slug", ""), c["name"]))
        pool += fs_candidates(http, c["name"])
        for alias in SEARCH_ALIASES.get(c["name"], []):
            pool += fs_candidates(http, alias)
        seen_ids = set()
        pool = [p for p in pool if not (p[0] in seen_ids or seen_ids.add(p[0]))]

        chosen, reason = None, "no candidate found"
        for cid, slug, _title in pool:
            ok, why = fs_verify(http, cid, slug, want_country, c["name"], roster)
            if ok:
                chosen, reason = (cid, slug), why
                break
            reason = why                       # keep the last real explanation

        row = {
            "league": c["league"],
            "country": c["country"],
            "club": c["name"],
            "sofa_id": c["sofa_id"],
            "fs_id": chosen[0] if chosen else "",
            "fs_slug": chosen[1] if chosen else "",
            "status": "verified" if chosen else "TODO",
            "checked_by": VERIFY_VERSION if chosen else "",
            "note": reason,
            "lookup": "" if chosen else f"{FS_BASE}/wyszukiwanie/?q={quote(c['name'])}",
        }
        rows.append(row)
        if chosen:
            verified += 1
        else:
            todo += 1
            print(f"    TODO {c['name']:<28} ({c['league']}) — {reason}")

    with OUT_CROSSWALK.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=REGISTER_COLS, extrasaction="ignore")
        w.writeheader()
        w.writerows(rows)

    print(f"  -> {OUT_CROSSWALK.name}: {len(rows)} clubs, "
          f"{verified} verified, {todo} need attention")
    if todo:
        print("     For each TODO row: open the lookup URL, click the club, and copy")
        print("     the two URL parts into fs_slug and fs_id. Re-run to verify them.")
        print("     Verified rows are never overwritten.")
    return rows


# ----------------------------------------------------------------------------
# Stage 3 — rosters (Sofascore) and positions (Flashscore)
# ----------------------------------------------------------------------------

def sofa_roster(http: Http, sofa_id: int) -> list[dict]:
    data = http.get_json(f"{SOFA}/team/{sofa_id}/players")
    out = []
    for entry in data.get("players", []):
        p = entry.get("player", entry)
        if not p.get("name"):
            continue
        # `name` is reliably natural order. firstName/lastName are NOT populated
        # consistently (sometimes the full name lands in firstName), so ignore them.
        out.append({
            "name": to_ascii(p["name"]),
            "sofa_pos": SOFA_POS.get(p.get("position", ""), ""),
        })
    return out


def fs_player_tokens(html: str) -> list[frozenset[str]]:
    """Every player name on a Flashscore squad page, as token sets."""
    out = []
    for m in re.finditer(
        r'<a[^>]+href="[^"]*/(?:zawodnik|player)/[^/"]+/([A-Za-z0-9]{6,10})/?"[^>]*>(.*?)</a>',
        html, re.S,
    ):
        name = strip_tags(m.group(2))
        name = re.split(r"\s{2,}|Kontuzja|Injury|Zawieszenie", name)[0].strip()
        tk = tokens(name)
        if tk:
            out.append(tk)
    return out


def fs_positions(http: Http, fs_id: str, fs_slug: str) -> dict[frozenset[str], str]:
    """
    Parse a Flashscore squad page into {name-token-set: position}.

    Keyed off the visible section headings rather than CSS classes, so a
    frontend reskin does not break it. The page concatenates every competition
    tab, so we take the union across tabs and resolve the rare disagreement by
    majority (ties go to the last block, which is the cumulative one).
    """
    slug = fs_slug.strip("/").split("/")[-1] if fs_slug else "team"
    url = f"{FS_BASE}/druzyna/{slug}/{fs_id}/sklad/"
    try:
        html = http.get(url, referer=FS_BASE, cache_key=f"fssquad_{fs_id}")
    except Exception as e:
        print(f"    Flashscore squad fetch failed for {fs_id}: {e}")
        return {}

    # Locate every heading and every player link, then assign links to whichever
    # heading most recently precedes them.
    marks: list[tuple[int, str]] = []
    for m in re.finditer(r">\s*([A-Za-zÀ-ž]+)\s*<", html):
        canon = FS_HEADINGS.get(to_ascii(m.group(1)).lower())
        if canon:
            marks.append((m.start(), canon))

    votes: dict[frozenset[str], Counter] = {}
    for m in re.finditer(
        r'<a[^>]+href="[^"]*/(?:zawodnik|player)/[^/"]+/([A-Za-z0-9]{6,10})/?"[^>]*>(.*?)</a>',
        html, re.S,
    ):
        name = strip_tags(m.group(2))
        # Trim injury/status text Flashscore appends inside the row.
        name = re.split(r"\s{2,}|Kontuzja|Injury|Zawieszenie", name)[0].strip()
        tk = tokens(name)
        if not tk:
            continue
        pos = ""
        for idx, canon in marks:
            if idx < m.start():
                pos = canon
            else:
                break
        if pos and pos != "__coach__":
            votes.setdefault(tk, Counter())[pos] += 1

    return {tk: c.most_common(1)[0][0] for tk, c in votes.items()}


def stage_squads(http: Http, crosswalk: list[dict]) -> list[dict]:
    records: list[dict] = []
    for i, club in enumerate(crosswalk, 1):
        club_name = club.get("club") or club.get("name", "")
        print(f"  [{i}/{len(crosswalk)}] {club_name}")
        try:
            roster = sofa_roster(http, int(club["sofa_id"]))
        except Exception as e:
            print(f"    !! roster failed: {e}")
            continue

        # Only a verified mapping is allowed to supply positions. An unproven
        # one is worse than none: it silently attributes another club's squad.
        use_fs = club.get("fs_id") and club.get("status", "verified") == "verified"
        fs_map = fs_positions(http, club["fs_id"], club.get("fs_slug", "")) if use_fs else {}
        fs_index = list(fs_map.items())

        for p in roster:
            tk = tokens(p["name"])
            pos, source = "", ""

            if tk in fs_map:                                  # exact token-set hit
                pos, source = fs_map[tk], "flashscore"
            else:
                # Near-miss: handles middle names present on only one side.
                # Every candidate above threshold is collected, not just the best,
                # because a single winner is only trustworthy if it is unopposed.
                cands = []
                for cand_tk, cand_pos in fs_index:
                    inter = tk & cand_tk
                    if not inter:
                        continue
                    subset = tk <= cand_tk or cand_tk <= tk
                    # A one-token overlap is only meaningful if that token is
                    # distinctive. Mononyms ("Ederson", "Fred", "Talisca") are a
                    # subset of any longer name sharing the token, so requiring
                    # two shared tokens stops them latching onto a namesake.
                    if len(inter) < 2 and not (subset and min(len(tk), len(cand_tk)) >= 2):
                        continue
                    score = len(inter) / max(len(tk | cand_tk), 1) + (0.35 if subset else 0)
                    if score >= 0.55:
                        cands.append((score, cand_tk, cand_pos))

                distinct = {c[1] for c in cands}
                if len(distinct) == 1:
                    pos, source = cands[0][2], "flashscore~fuzzy"
                elif len(distinct) > 1:
                    # Genuinely ambiguous. Never guess — fall through to the
                    # Sofascore value and flag the row for review.
                    source = "ambiguous"

            if not pos:
                pos = p["sofa_pos"]
                source = source or "sofascore"
                if source == "ambiguous":
                    source = "ambiguous->sofascore"

            if pos == "Goalkeeper" or not pos:
                continue                                       # goalkeepers excluded

            records.append({
                "Name": p["name"],
                "Position": pos,
                "Club": club_name,
                "Country of the club": club["country"],
                "_league": club["league"],
                "_source": source,
            })

    OUT_SQUADS.write_text(json.dumps(records, indent=2, ensure_ascii=False), encoding="utf-8")
    print(f"  -> {OUT_SQUADS.name} ({len(records)} players)")
    return records


# ----------------------------------------------------------------------------
# Stage 4 — output
# ----------------------------------------------------------------------------

def stage_build(records: list[dict]) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font
    from openpyxl.utils import get_column_letter

    cols = ["Name", "Position", "Club", "Country of the club"]
    widths = [30, 14, 26, 22]

    # Translate at write time only. Both maps fall back to the original value,
    # so this is idempotent and a missing entry degrades to English rather
    # than to a blank cell.
    def localise(rec: dict) -> dict:
        if not POLISH_OUTPUT:
            return rec
        out = dict(rec)
        out["Position"] = POSITION_PL.get(rec["Position"], rec["Position"])
        out["Country of the club"] = COUNTRY_PL.get(
            rec["Country of the club"], rec["Country of the club"])
        return out

    records = [localise(r) for r in records]
    header = [HEADERS_PL.get(c, c) if POLISH_OUTPUT else c for c in cols]

    wb = Workbook()
    ws = wb.active
    ws.title = "Skład" if POLISH_OUTPUT else "Squads"
    ws.append(header)
    for r in sorted(records, key=lambda x: (x["_league"], x["Club"], x["Name"])):
        ws.append([r[c] for c in cols])

    for c in ws[1]:
        c.font = Font(bold=True)
        c.alignment = Alignment(horizontal="left")
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{ws.max_row}"
    wb.save(OUT_XLSX)

    # Diagnostic only — not the deliverable. Lists rows whose position did not
    # come cleanly from Flashscore, so mistakes are visible instead of silent.
    import csv
    rcols = cols + ["_league", "_source"]
    flagged = [r for r in records if r["_source"] != "flashscore"]
    with OUT_REVIEW.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=rcols)
        w.writeheader()
        w.writerows(flagged)

    by_source = Counter(r["_source"] for r in records)
    total = len(records) or 1
    print(f"\n  -> {OUT_XLSX.name}: {len(records)} players")
    print(f"  -> {OUT_REVIEW.name}: {len(flagged)} rows needing a look\n")
    print("  Position provenance:")
    for src, n in by_source.most_common():
        print(f"    {src:<20} {n:>5}  ({n / total:5.1%})")


# ----------------------------------------------------------------------------

def load(path: Path) -> Any:
    if not path.exists():
        sys.exit(f"Missing {path.name}. Run the earlier stage first.")
    return json.loads(path.read_text(encoding="utf-8"))


def read_crosswalk() -> list[dict]:
    import csv
    if not OUT_CROSSWALK.exists():
        sys.exit(f"Missing {OUT_CROSSWALK.name}. Run:  python3 squad_scraper.py crosswalk")
    with OUT_CROSSWALK.open(encoding="utf-8") as f:
        return [r for r in csv.DictReader(f)]


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("stage", choices=["clubs", "crosswalk", "squads", "build", "all"])
    ap.add_argument("--pilot", action="store_true", help="2 clubs from one league only")
    ap.add_argument("--refresh", action="store_true",
                    help="ignore cached responses and re-fetch everything")
    ap.add_argument("--max-age", type=float, default=24.0, metavar="HOURS",
                    help="reuse cached responses younger than this (default: 24)")
    ap.add_argument("--cache-dir", type=Path, default=None, metavar="PATH",
                    help=f"where to cache responses (default: {default_cache_dir()})")
    ap.add_argument("--quiet", action="store_true")
    args = ap.parse_args()

    max_age = 0.0 if args.refresh else args.max_age
    http = Http(verbose=not args.quiet, max_age_hours=max_age, cache_dir=args.cache_dir)
    if max_age == 0:
        print("\nCache bypassed — fetching everything fresh.")
    t0 = time.time()

    if args.stage in ("clubs", "all"):
        print("\n[1/4] Discovering clubs")
        clubs = stage_clubs(http, args.pilot)
    else:
        clubs = load(OUT_CLUBS) if OUT_CLUBS.exists() else []

    if args.stage in ("crosswalk", "all"):
        print("\n[2/4] Building Sofascore <-> Flashscore crosswalk")
        crosswalk = stage_crosswalk(http, clubs or load(OUT_CLUBS))
    elif args.stage in ("squads", "build"):
        crosswalk = read_crosswalk()
    else:
        crosswalk = []

    if args.stage in ("squads", "all"):
        print("\n[3/4] Pulling rosters and positions")
        records = stage_squads(http, crosswalk)
    else:
        records = load(OUT_SQUADS) if OUT_SQUADS.exists() else []

    if args.stage in ("build", "all"):
        print("\n[4/4] Writing output")
        stage_build(records or load(OUT_SQUADS))

    print(f"\nDone in {time.time() - t0:.0f}s "
          f"({http.fetches} fetched, {http.hits} reused from cache)")
    print(f"Cache: {http.cache} ({dir_size_mb(http.cache):.0f}MB) — safe to delete")
    if http.fetches == 0 and args.stage != "build":
        print("Every response came from cache — this data is not fresh. "
              "Use --refresh to force a live scrape.")


if __name__ == "__main__":
    main()
